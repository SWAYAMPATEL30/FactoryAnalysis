"""Stage 7 -- Excel Generation.

Copies the real target template (never rebuilds it) and writes one row per
MostRow into 'MOST Analysis'. Formula columns (Q, S, W, AC) are written as
live Excel formulas -- translated per-row from the template's own formula
text -- so the workbook keeps recalculating if a reviewer edits a cell by
hand in Stage 8. Only the taxonomy-bucket columns (Y/Z/AA/AB) intentionally
use a classification-name lookup instead of the template's numeric-ref-range
IF()s -- see the note below.

Sheet 2 ('VA SVA NVA Summary') is left completely untouched: its SUMIF/pivot
formulas already reference whole columns on Sheet 1, so they keep working
unmodified as rows are added.

Two latent bugs in the original workbook's Y/Z/AA/AB formulas, found while
building this (neither is exercised by the example workbook's own 26 rows,
so they were invisible until now):
  1. Ref 0 (Noise) satisfies "V<23" and is silently counted as NVA time.
  2. Ref 23 ("Manual testing", taxonomy classification NVA) satisfies
     none of Y/Z/AA/AB's range conditions and is silently dropped
     from every bucket.
Both stem from bucketing by numeric ref ranges instead of the taxonomy's own
classification field. Since the taxonomy is required to be versioned/editable
(refs can move -- ref 48/50 already did in this build), this writer buckets
new rows via VLOOKUP against the taxonomy's classification column instead,
which is immune to both bugs and to any future renumbering.
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from app.config.most_tables import load_most_tables
from app.models.schemas import MostRow

SHEET_NAME = "MOST Analysis"
FIRST_DATA_ROW = 6
MAX_CLEAR_COL = 45  # generous headroom above AM (col 39)

# Extra traceability & activity columns, appended after the template's own AC (col 29).
TRACE_HEADERS = {
    "AD": "Source Video",
    "AE": "Segment Start (s)",
    "AF": "Segment End (s)",
    "AG": "Segmentation Model",
    "AH": "Classification Model",
    "AI": "Confidence",
    "AJ": "Human Corrected",
    "AK": "Activity & Movement Details",
    "AL": "Activity Duration (sec)",
    "AM": "Activity Timeline",
    "AN": "Elemental Description",
}


def _clear_existing_data_rows(ws) -> None:
    from openpyxl.utils import get_column_letter

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, MAX_CLEAR_COL + 1):
            ws[f"{get_column_letter(c)}{r}"] = None


def _write_trace_headers(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Set explicit trace header texts first
    for col, header in TRACE_HEADERS.items():
        ws[f"{col}5"] = header

    # Format ALL header cells in row 5 from column A (1) to AN (40)
    for c in range(1, 41):
        col_letter = get_column_letter(c)
        cell = ws[f"{col_letter}5"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border


def _create_timeline_chart_sheet(wb, rows: list[MostRow]) -> None:
    """Creates a dedicated 'Activity Timeline Chart' worksheet with a visual Gantt chart
    displaying the chronological distribution, durations, and category breakdowns of all activities."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    chart_sheet_name = "Activity Timeline Chart"
    if chart_sheet_name in wb.sheetnames:
        del wb[chart_sheet_name]

    ws_chart = wb.create_sheet(title=chart_sheet_name)
    ws_chart.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_chart.merge_cells("A1:G1")
    title_cell = ws_chart["A1"]
    title_cell.value = "Activity Timeline & Distribution Chart"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_chart.row_dimensions[1].height = 35

    # Headers
    headers = [
        "Activity Description",
        "Start Time (s)",
        "Duration (s)",
        "End Time (s)",
        "Category",
        "Movement State",
        "Machine State",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws_chart.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_chart.row_dimensions[3].height = 25

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    start_row = 4
    for i, r in enumerate(rows):
        row_num = start_row + i
        ws_chart.cell(row=row_num, column=1, value=r.elemental_description)
        ws_chart.cell(row=row_num, column=2, value=r.t_start_sec)
        ws_chart.cell(row=row_num, column=3, value=r.activity_duration_sec)
        ws_chart.cell(row=row_num, column=4, value=r.t_end_sec)
        ws_chart.cell(row=row_num, column=5, value=r.category)

        # Parse movement state & machine state from details string
        mov_state = "MOVE"
        mac_state = "IDLE"
        if "(" in r.activity_movement_details and ")" in r.activity_movement_details:
            mov_state = r.activity_movement_details.split("(")[1].split(")")[0]
        if "Machine:" in r.activity_movement_details:
            mac_state = r.activity_movement_details.split("Machine:")[1].strip()

        ws_chart.cell(row=row_num, column=6, value=mov_state)
        ws_chart.cell(row=row_num, column=7, value=mac_state)

        for col in range(1, 8):
            c = ws_chart.cell(row=row_num, column=col)
            c.border = thin_border
            if col in (2, 3, 4):
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")

    num_rows = len(rows)
    end_row = start_row + num_rows - 1

    # Create Horizontal Stacked Bar Chart for Gantt Timeline representation
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Activity Execution Timeline (Gantt Chart)"
    chart.height = max(10, num_rows * 0.8)
    chart.width = 22

    # Series: Start Time (invisible offset) and Duration (active bar)
    data = Reference(ws_chart, min_col=2, min_row=3, max_col=3, max_row=end_row)
    cats = Reference(ws_chart, min_col=1, min_row=4, max_row=end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.x_axis.title = "Timeline (Seconds)"
    chart.y_axis.title = "Activities"
    chart.legend = None

    # Embed chart beside table
    ws_chart.add_chart(chart, "I3")

    # Column widths
    ws_chart.column_dimensions["A"].width = 38
    ws_chart.column_dimensions["B"].width = 15
    ws_chart.column_dimensions["C"].width = 15
    ws_chart.column_dimensions["D"].width = 15
    ws_chart.column_dimensions["E"].width = 25
    ws_chart.column_dimensions["F"].width = 18
    ws_chart.column_dimensions["G"].width = 18


def _append_pareto_section(ws, rows: list) -> None:
    """Appends a Pareto analysis section below the existing Gantt chart data in
    the 'Activity Timeline Chart' sheet (Tab 3).  Writes a data table sorted
    descending by TMU, highlights the top row, then embeds a native combo
    BarChart + LineChart (secondary axis) so the chart is fully editable inside
    Excel."""
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    TMU_TO_SEC = 0.036
    total_tmu = sum(r.tmu for r in rows) if rows else 1
    if total_tmu == 0:
        total_tmu = 1

    # Sort descending by TMU
    sorted_rows = sorted(rows, key=lambda r: r.tmu, reverse=True)

    # Find the last used row in the sheet
    last_used = ws.max_row
    section_start = last_used + 4  # leave 3 blank rows as visual separator

    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Section title banner
    ws.merge_cells(
        start_row=section_start, start_column=1,
        end_row=section_start, end_column=5
    )
    title_cell = ws.cell(row=section_start, column=1)
    title_cell.value = "PARETO ANALYSIS \u2014 Time by Activity (sorted descending)"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[section_start].height = 30

    # Sub-header row
    header_row = section_start + 1
    headers = ["Activity", "TMU", "Time (sec)", "% of Total", "Cumulative %"]
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[header_row].height = 20

    # Data rows
    data_start = header_row + 1
    cumulative_pct = 0.0
    for i, r in enumerate(sorted_rows):
        row_num = data_start + i
        own_sec = r.tmu * TMU_TO_SEC
        own_pct = (r.tmu / total_tmu) * 100
        cumulative_pct += own_pct

        values = [
            r.elemental_description or f"Activity {r.s_no}",
            r.tmu,
            round(own_sec, 3),
            round(own_pct, 2),
            round(min(cumulative_pct, 100.0), 2),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00"

            # Highlight top row (highest TMU) in amber
            if i == 0:
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, bold=True)
            else:
                cell.font = Font(name="Calibri", size=10)

    data_end = data_start + len(sorted_rows) - 1

    # Column widths
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 40)
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    # ── Native combo chart: BarChart (time sec) + LineChart (cumulative %) ──
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Pareto — Activity Time & Cumulative %"
    bar.y_axis.title = "Time (seconds)"
    bar.x_axis.title = "Activity"
    bar.grouping = "clustered"
    bar.height = 14
    bar.width = 24

    # Bar data: Time (sec) column (col 3)
    bar_data = Reference(ws, min_col=3, min_row=header_row, max_row=data_end)
    bar_cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)

    # Line chart: Cumulative % column (col 5)
    line = LineChart()
    line.y_axis.axId = 200
    line.y_axis.title = "Cumulative %"
    line.y_axis.crosses = "max"  # render on right
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 100

    line_data = Reference(ws, min_col=5, min_row=header_row, max_row=data_end)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(bar_cats)

    # Combine: bar absorbs the line on secondary axis
    bar += line

    chart_anchor_row = data_start
    ws.add_chart(bar, f"G{chart_anchor_row}")


def _write_insights_tab(wb, insights) -> None:
    """Creates a dedicated 'Improvement Insights' worksheet (Tab 4) containing
    structured recommendations for Sections A, B, C, and D."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sheet_name = "Improvement Insights"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True

    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title Banner
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "AI CYCLE-IMPROVEMENT INSIGHTS REPORT"
    title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Section A: Bottleneck
    ws.merge_cells("A3:F3")
    sec_a = ws["A3"]
    sec_a.value = "A. BOTTLENECK IDENTIFICATION (Primary Time Consumer)"
    sec_a.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec_a.fill = PatternFill(start_color="C8452C", end_color="C8452C", fill_type="solid")
    ws.row_dimensions[3].height = 25

    b = insights.bottleneck
    ws["A4"] = "Activity Name:"
    ws["B4"] = b.activity_name
    ws["A5"] = "Cycle Share:"
    ws["B5"] = f"{b.time_sec:.1f}s ({b.pct_of_cycle:.1f}% of cycle time) - {b.tmu:.0f} TMU"
    ws["A6"] = "Root Cause Analysis:"
    ws["B6"] = b.reason

    for r_num in range(4, 7):
        ws[f"A{r_num}"].font = Font(name="Calibri", size=10, bold=True)

    # Section B: Elimination Candidates
    ws.merge_cells("A8:F8")
    sec_b = ws["A8"]
    sec_b.value = "B. NON-VALUE-ADD ELIMINATION CANDIDATES"
    sec_b.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec_b.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    ws.row_dimensions[8].height = 25

    headers_b = ["#", "Activity Description", "Current Time (s)", "Waste Category", "Data-Grounded Reason", "Potential Saving (s)"]
    for c_idx, h in enumerate(headers_b, 1):
        cell = ws.cell(row=9, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row_cursor = 10
    for cand in insights.elimination_candidates:
        ws.cell(row=row_cursor, column=1, value=cand.s_no).border = thin
        ws.cell(row=row_cursor, column=2, value=cand.activity_name).border = thin
        ws.cell(row=row_cursor, column=3, value=cand.current_time_sec).border = thin
        ws.cell(row=row_cursor, column=4, value=cand.waste_type).border = thin
        ws.cell(row=row_cursor, column=5, value=cand.reason).border = thin
        ws.cell(row=row_cursor, column=6, value=cand.potential_saving_sec).border = thin
        row_cursor += 1

    # Section C: Equipment & Method Upgrades
    row_cursor += 1
    ws.merge_cells(f"A{row_cursor}:G{row_cursor}")
    sec_c = ws[f"A{row_cursor}"]
    sec_c.value = "C. EQUIPMENT & METHOD UPGRADE SUGGESTIONS"
    sec_c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec_c.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    row_cursor += 1

    headers_c = ["#", "Activity Description", "Current Method / Tool", "Suggested Upgrade", "Projected Time (s)", "Disclaimer", "Web Search Verification"]
    for c_idx, h in enumerate(headers_c, 1):
        cell = ws.cell(row=row_cursor, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row_cursor += 1

    for eq in insights.equipment_upgrades:
        ws.cell(row=row_cursor, column=1, value=eq.s_no).border = thin
        ws.cell(row=row_cursor, column=2, value=eq.activity_name).border = thin
        ws.cell(row=row_cursor, column=3, value=eq.current_tool_or_method).border = thin
        ws.cell(row=row_cursor, column=4, value=eq.suggested_upgrade).border = thin
        ws.cell(row=row_cursor, column=5, value=eq.projected_time_sec).border = thin
        disc_cell = ws.cell(row=row_cursor, column=6, value=eq.disclaimer)
        disc_cell.border = thin
        disc_cell.font = Font(name="Calibri", size=9, italic=True, color="C8452C")

        link_cell = ws.cell(row=row_cursor, column=7)
        link_cell.border = thin
        if getattr(eq, "search_url", None):
            link_cell.value = f'=HYPERLINK("{eq.search_url}", "🔍 Search Product Specs")'
            link_cell.font = Font(name="Calibri", size=9, color="0000FF", underline="single")
        else:
            link_cell.value = "N/A"
        row_cursor += 1

    # Section D: Summary
    row_cursor += 1
    ws.merge_cells(f"A{row_cursor}:G{row_cursor}")
    sec_d = ws[f"A{row_cursor}"]
    sec_d.value = "D. PROJECTED NEW CYCLE TIME SUMMARY"
    sec_d.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec_d.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    row_cursor += 1

    s = insights.projected_summary
    ws[f"A{row_cursor}"] = "Baseline Cycle Time:"
    ws[f"B{row_cursor}"] = f"{s.current_cycle_sec:.2f}s ({s.current_tmu:.0f} TMU)"
    row_cursor += 1
    ws[f"A{row_cursor}"] = "Projected Cycle Time:"
    ws[f"B{row_cursor}"] = f"{s.projected_cycle_sec:.2f}s ({s.projected_tmu:.0f} TMU)"
    row_cursor += 1
    ws[f"A{row_cursor}"] = "Total Time Savings:"
    ws[f"B{row_cursor}"] = f"-{s.total_saving_sec:.2f}s (-{s.pct_reduction:.1f}%)"
    row_cursor += 1
    ws[f"A{row_cursor}"] = "Disclaimer:"
    disc_summary = ws[f"B{row_cursor}"]
    disc_summary.value = s.disclaimer
    disc_summary.font = Font(name="Calibri", size=10, italic=True, color="C8452C")

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 28


def write_most_analysis_workbook(
    rows: list[MostRow],
    template_path: Path,
    output_path: Path,
    activity_description: str,
    insights: object | None = None,
) -> Path:
    if not rows:
        raise ValueError("no rows to write")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copyfile(template_path, output_path)
    except PermissionError:
        import time
        output_path = output_path.with_name(f"{output_path.stem}_{int(time.time())}{output_path.suffix}")
        shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb[SHEET_NAME]

    _clear_existing_data_rows(ws)
    _write_trace_headers(ws)

    tables = load_most_tables()
    q_template = "=IF(E6=$E$1,($F$1&F6&$G$1&G6&$H$1&H6&$I$1&I6&$J$1&J6&$K$1&K6&$L$1&L6),IF(E6=$E$2,($F$2&F6&$G$2&G6&$H$2&H6&$I$2&I6&$J$2&J6&$K$2&K6&$L$2&L6),IF(E6=$E$3,($F$3&F6&$G$3&G6&$H$3&H6&$I$3&I6&$J$3&J6&$K$3&K6&$L$3&L6&$M$3&M6&$N$3&N6&$O$3&O6&$P$3&P6),IF(E6=$E$4,F6&\"SEC\"))))"
    s_template = "=IF(E6=$E$4,R6*F6/0.036,(SUM(F6:P6)*10*R6))"
    ac_template = "=VLOOKUP(V6,'VA SVA NVA Summary'!A:E,3,0)"

    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        model = tables.sequence_models[row.data_card]

        ws[f"A{r}"] = row.s_no
        ws[f"B{r}"] = row.station_no
        ws[f"C{r}"] = row.activity_no
        ws[f"D{r}"] = activity_description
        ws[f"E{r}"] = row.data_card
        for col, value in zip(model.columns, row.param_values):
            ws[f"{col}{r}"] = value

        ws[f"Q{r}"] = Translator(q_template, origin="Q6").translate_formula(f"Q{r}")
        ws[f"R{r}"] = row.freq
        ws[f"S{r}"] = Translator(s_template, origin="S6").translate_formula(f"S{r}")
        ws[f"T{r}"] = row.elemental_description
        ws[f"U{r}"] = row.operator
        ws[f"V{r}"] = row.muda_ref
        ws[f"W{r}"] = f"=S{r}*0.036"
        ws[f"X{r}"] = row.online_offline_mode

        classification_lookup = f"VLOOKUP(V{r},'VA SVA NVA Summary'!$A:$E,2,0)"
        ws[f"Y{r}"] = f'=IF({classification_lookup}="VA",W{r},0)'
        ws[f"Z{r}"] = f'=IF({classification_lookup}="NVA-N",W{r},0)'
        ws[f"AA{r}"] = f'=IF({classification_lookup}="SVA",W{r},0)'
        ws[f"AB{r}"] = f'=IF({classification_lookup}="NVA",W{r},0)'
        ws[f"AC{r}"] = Translator(ac_template, origin="AC6").translate_formula(f"AC{r}")

        ws[f"AD{r}"] = row.source_video_uri
        ws[f"AE{r}"] = row.t_start_sec
        ws[f"AF{r}"] = row.t_end_sec
        ws[f"AG{r}"] = row.segment_model_version
        ws[f"AH{r}"] = row.classification_model_version
        ws[f"AI{r}"] = row.confidence
        ws[f"AJ{r}"] = "YES" if row.human_corrected else "NO"
        ws[f"AK{r}"] = row.activity_movement_details
        ws[f"AL{r}"] = row.activity_duration_sec
        ws[f"AM{r}"] = row.activity_timeline
        ws[f"AN{r}"] = row.uppercase_elemental_description

        # Apply clean thin borders across ALL data cells (columns A to AN)
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter

        thin_data_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for c in range(1, 41):
            col_letter = get_column_letter(c)
            ws[f"{col_letter}{r}"].border = thin_data_border

    # Label column W header so it's human-readable in Excel
    ws["W5"] = "TMU Time (sec)"

    last_row = FIRST_DATA_ROW + len(rows) - 1
    ws["D6"] = activity_description
    ws["W4"] = f"=SUM(W{FIRST_DATA_ROW}:W{last_row})"

    # Generate dedicated Activity Timeline Chart worksheet tab (Tab 3)
    _create_timeline_chart_sheet(wb, rows)

    # Append Pareto section below the Gantt chart on the same Tab 3
    ws_chart = wb["Activity Timeline Chart"]
    _append_pareto_section(ws_chart, rows)

    # Write Tab 4 if insights are provided
    if insights is not None:
        _write_insights_tab(wb, insights)

    wb.save(output_path)
    return output_path
